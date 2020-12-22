# regionalization.py
# Created by Hannah Chan
# Date: November 2020

#####################################################################
"""
This file takes in an excel workbook and produces a regionalized map
    and saves it under another workbook.

Input:
Before beginning, input excel file must have the following:
    Sheet1 = "[area]"              --> ex. CAN for canada, copy paste the .asc file
    Others = "[abbreviation/name]" --> copy pasted .asc file 
    **optional:
    list = "list" --> list of all region numbers and names

Output:
    Sheet1 = "map"          --> regionalized map
    Sheet2 = "legend"       --> region numbers and names
    Sheet3 = "overlaps"     --> cells with overlap between regions


openpyxl is used to control excel

"""
##############################################################################
import sys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.formatting.rule import ColorScaleRule
import re
from string import digits
import csv
import os
from user_inputs import *


##############################################################################
# functions for regionalization
##############################################################################



# returns a dictionary with info from header of each file 
# (first 2 columns and first 6 rows of a worksheet)
def get_file_header(ws):
    ncols = ws['B1'].value
    nrows = ws['B2'].value
    xllcorner = ws['B3'].value # bottom left corner
    yllcorner = ws['B4'].value # bottom left corner (column)
    cellsize = ws['B5'].value
    nodata_value = ws['B6'].value
    file_header = {
        "ncols"         :   ncols,
        "nrows"         :   nrows,
        "xllcorner"     :   xllcorner,
        "yllcorner"     :   yllcorner,
        "cellsize"      :   cellsize,
        "nodata_value"  :   nodata_value
    }
    return file_header


# returns a dictionary with basic information about top left and bottom left corners
def get_area_cell_info(area_header):

    # bottom left col coord (combined region)   --> used later for individual regions
    bottom_left_coord_to_cell_col = int(round(area_header["xllcorner"] /
                                        area_header["cellsize"])) 
    # bottom left row coord (combined region)   --> used later for individual regions
    bottom_left_coord_to_cell_row = int(round(area_header["yllcorner"] /
                                        area_header["cellsize"]))

    bottom_left_col = 1 + num_extra_left_cols
    bottom_left_row = area_header["nrows"] + num_extra_top_rows
    top_left_col = bottom_left_col
    top_left_row = bottom_left_row - area_header["nrows"] + 1 # looking for first row

    area_cell_info = {
        "bottom_left_coord_to_cell_col" :   bottom_left_coord_to_cell_col,
        "bottom_left_coord_to_cell_row" :   bottom_left_coord_to_cell_row,
        "bottom_left_col"               :   bottom_left_col,
        "bottom_left_row"               :   bottom_left_row,
        "top_left_col"                  :   top_left_col,
        "top_left_row"                  :   top_left_row
    } 
    return area_cell_info


# returns a dictionary with basic information about top left and bottom left corners
# note: rows and columns are for placement in area sheet (not individual region sheets)
def get_region_cell_info(region_header, area_header, area_cell_info):

    bottom_left_coord_to_cell_col = int(round(region_header["xllcorner"] /
                                        region_header["cellsize"])) 
    bottom_left_coord_to_cell_row = int(round(region_header["yllcorner"] /
                                        region_header["cellsize"]))

    # calculate the top left cell  
    bottom_left_col = (area_cell_info["bottom_left_col"] + 
                        (bottom_left_coord_to_cell_col - 
                        area_cell_info["bottom_left_coord_to_cell_col"]))
    
    bottom_left_row = (area_cell_info["bottom_left_row"] - 
                        (bottom_left_coord_to_cell_row - 
                        area_cell_info["bottom_left_coord_to_cell_row"]))
    
    top_left_col = bottom_left_col
    top_left_row = bottom_left_row - region_header["nrows"] + 1 # +1 because looking for first row

    region_cell_info = {
        "bottom_left_coord_to_cell_col" :   bottom_left_coord_to_cell_col,
        "bottom_left_coord_to_cell_row" :   bottom_left_coord_to_cell_row,
        "bottom_left_col"               :   bottom_left_col,
        "bottom_left_row"               :   bottom_left_row,
        "top_left_col"                  :   top_left_col,
        "top_left_row"                  :   top_left_row
    } 
    
    return region_cell_info


# this function gets the address of a cell when receiving row and col as numbers
def get_cell_address(col, row):
    col_letter = get_column_letter(col)
    address = col_letter + str(row)
    return address

# this function adds cell with overlap
def add_overlap(overlaps, row, col):
    overlaps.append(get_cell_address(col, row))
    return

# this function sets values in map_ws to the value of change_to_num, according to region_ws
# does not include blanks or nodatavalues 
def set_cells_to_num_except_blanks(map_ws, region_ws, region_header, region_cell_info, change_to_num, overlaps):
    ncols = region_header["ncols"]
    nrows = region_header["nrows"]
    nodata_value = region_header["nodata_value"]
    region_top_left_row = region_cell_info["top_left_row"]
    region_top_left_col = region_cell_info["top_left_col"]
    
    for row in range(nrows):
            for col in range(ncols):
                cell_val = region_ws.cell(row=row + num_extra_top_rows + 1, 
                                            column=col + num_extra_left_cols + 1).value
                
                # check to see if there will be overlap
                map_ws_current_row = region_top_left_row + row
                map_ws_current_col = region_top_left_col + col
                map_ws_current_val = map_ws.cell(row=map_ws_current_row, 
                                    column=map_ws_current_col).value

                if cell_val not in (nodata_value, None, ""):
                    # check to see if there will be overlap
                    if map_ws_current_val not in (None, ""):
                        add_overlap(overlaps, map_ws_current_row, map_ws_current_col)

                    # change value on map_ws
                    map_ws.cell(row=map_ws_current_row, 
                                    column=map_ws_current_col).value = change_to_num

# this function takes in information from each region
# and adds that information to map_ws
def each_region(wb, regions, map_ws, area_header, area_cell_info, overlaps, xlFilename):
    for num in regions:
        region = regions[num]

        # if given region does not exist in workbook
        if not region in wb.sheetnames:
            print("ERROR: A worksheet with region name '" + 
                region + "' does not exist in workbook '" + xlFilename +"'")
            continue
        
        region_ws = wb[region]
        region_header = get_file_header(region_ws)
        region_cell_info = get_region_cell_info(region_header, area_header, area_cell_info)
        
        set_cells_to_num_except_blanks(map_ws, region_ws, region_header, region_cell_info, num, overlaps)

        print("Now finished region: " + regions[num])
    
    return


# this function sets range of 'A1':'B6' to be equal 
# between sheets map_ws & area_ws
def set_headers_equal(map_ws, area_ws):
    # set header in map as the same as header in area
    cellsToCopy = []
    for letter in ['A','B']:
        for num in range(1,7):
            cellsToCopy.append(letter + str(num))

    for val in cellsToCopy:
        map_ws[val].value = area_ws[val].value

# this function sets blanks in map_ws that are within borders to the no data value
# information from the area sheet is used 
def set_blanks_to_nodata(map_ws, area_header, area_cell_info):
    ncols = area_header["ncols"]
    nrows = area_header["nrows"]
    nodata_value = area_header["nodata_value"]
    area_top_left_row = area_cell_info["top_left_row"]
    area_top_left_col = area_cell_info["top_left_col"]
    
    for row in range(nrows):
            for col in range(ncols):
                cell_val = map_ws.cell(row=row + num_extra_top_rows + 1, 
                                            column=col + num_extra_left_cols + 1).value
                if (cell_val == "") or (cell_val == None):
                    map_ws.cell(row=area_top_left_row + row, 
                                    column=area_top_left_col + col).value = nodata_value

# this function sets column_width to user-defined variable
# taken from https://stackoverflow.com/a/60801712
def set_column_width(map_ws):
    col_width = 2
    dim_holder = DimensionHolder(worksheet=map_ws)

    for col in range(map_ws.min_column, map_ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(map_ws, min=col, max=col, width=col_width)

    map_ws.column_dimensions = dim_holder

# this function sets a color scale for range within borders of map_ws
# colours and scales set by user above
def set_color_scale(map_ws, area_header, area_cell_info):
    # the following is for the color scale
    color_start_value = 00 # percentage (between 0-100)
    color_start_value_color = 'ED5F49' # light red
    color_mid_value = 60 # percentage (between 0-100)
    color_mid_value_color = 'CEE740' # yellow
    color_end_value = 100 # percentage (between 0-100)
    color_end_value_color = '22910C' # green


    colorscale_rule = ColorScaleRule(start_type='percentile', start_value=color_start_value, start_color=color_start_value_color,
                          mid_type='percentile', mid_value=color_mid_value, mid_color=color_mid_value_color,
                          end_type='percentile', end_value=color_end_value, end_color=color_end_value_color)
    
    top_left_col = area_cell_info['top_left_col']
    top_left_row = area_cell_info['top_left_row']
    bottom_right_col = area_cell_info['bottom_left_col'] + (area_header['ncols'] - 1)
    bottom_right_row = area_cell_info['bottom_left_row']

    top_left_cell_in_range = get_cell_address(top_left_col, top_left_row)
    bottom_right_cell_in_range = get_cell_address(bottom_right_col, bottom_right_row)
    range_to_format = top_left_cell_in_range + ":" + bottom_right_cell_in_range
    
    map_ws.conditional_formatting.add(range_to_format, colorscale_rule)


# this function formats map_ws
def format_map_ws(map_ws, area_ws, area_header, area_cell_info, format_map):
    # set header of map_ws to be the same as area_ws
    set_headers_equal(map_ws, area_ws)
    # set to nodata value (of area) if blank
    set_blanks_to_nodata(map_ws, area_header, area_cell_info)

    if format_map:
        # set column width
        set_column_width(map_ws)
        # conditional format cells
        set_color_scale(map_ws, area_header, area_cell_info)

# this function prints the region names and numbers into legend_ws
def create_legend(legend_ws, regions):
    row = 1
    col1 = 1
    col2 = 2

    legend_ws.cell(row=row, column=col1).value = "region number" 
    legend_ws.cell(row=row, column=col2).value = "region abbreviation"

    for num in regions:
        row += 1
        legend_ws.cell(row=row, column=col1).value = num 
        legend_ws.cell(row=row, column=col2).value = regions[num]

# this function prints the region names and numbers into legend_ws
def print_overlaps(overlaps_ws, overlaps):
    row = 1
    col = 1

    overlaps_ws.cell(row=row, column=col).value = "list of cells with overlaps" 

    for cell in overlaps:
        row += 1
        overlaps_ws.cell(row=row, column=col).value = cell


# this function saves the sheets 'map' and 'legend' and 'overlaps' into new workbook
def save_files(wb, save_wb_name, save_csv_names, format_map):
    directory = 'Outputs/'
    if not os.path.exists(directory):
        os.makedirs(directory)
    # save sheets into csv file
    # using code inspired by https://stackoverflow.com/a/10803229
    save_sheets = ['map', 'legend', 'overlaps']
    for ws_name in save_sheets:
        print("Now saving " + ws_name + " to: " + save_csv_names[ws_name])

        with open(directory + save_csv_names[ws_name],'w', newline='') as file:
            writer = csv.writer(file)
            for row in wb[ws_name].rows:
                writer.writerow([cell.value for cell in row])
    
    # save only sheet 'map' by removing other sheets
    # code copied from https://stackoverflow.com/a/46237894
    if format_map:
        sheets = wb.sheetnames
        for ws_name in sheets:
            if (ws_name != 'map'):
                wb.remove(wb[ws_name])

        print("Now saving formatted map to workbook: " + save_wb_name)
        wb.save(directory + save_wb_name)
        print("Everything has been saved")




##############################################################################
# global variables
##############################################################################
num_extra_top_rows = 6 # the top of every asc file has 6 extra rows
num_extra_left_cols = 1 # the left of every asc file has 1 extra column
   


##############################################################################
# main script
##############################################################################



def main():
    line_end = "===================="
    line_begin = "\n" + line_end

    xlFilename, wb, save_csv_names, format_map, save_wb_name, area_name, regions = define_all_variables()
   
    # create sheets in workbook for regionalized map, legend, and overlaps
    map_ws = wb.create_sheet("map")
    legend_ws = wb.create_sheet("legend")
    overlaps_ws = wb.create_sheet("overlaps")
    
    # area worksheet
    area_ws = wb[area_name]
    area_header = get_file_header(area_ws)
    area_cell_info = get_area_cell_info(area_header)

    # list to keep track of cells with overlaps
    overlaps = []

    # go through each region and add to map_ws
    print(line_begin, "Making the regionalization map", line_end)
    each_region(wb, regions, map_ws, area_header, area_cell_info, overlaps, xlFilename)   

    # format map worksheet to have no data value, color scale, and smaller column width
    format_map_ws(map_ws, area_ws, area_header, area_cell_info, format_map)
    
    # create legend
    create_legend(legend_ws, regions)

    #print overlaps
    print_overlaps(overlaps_ws, overlaps)

    print(line_begin, "Saving files", line_end)
    # save workbook
    save_files(wb, save_wb_name, save_csv_names, format_map)
    
    return


main()


