
from openpyxl import load_workbook

##############################################################################
# functions for interaction with user
##############################################################################


def y_or_n(question):
    while True:
        ans = input(question + " (y/n): ")
        if ans in ('y', 'Y'):
            return True
        elif ans in ('n', 'N'):
            return False
        else:
            print("Your answer '" + ans + "' was invalid.", 
                  "Please choose 'y' or 'n'.")


def answer(question):
    while True:
        ans = input(question + ": ")
        check = input('Please confirm your answer: "' + ans + '" (y/n): ')
        if check in ('y', 'Y'):
            return ans
        else:
            print("Answer was not confirmed. Please enter answer again.") 

def print_dict(dc):
    for key in dc:
        print(key, ':', dc[key])

def print_regions(regions):
    if regions == {}:
        return
    print("The following are your current regions:")
    print("-----------------------\n",
            "    regions:    ", 
            "\n-----------------------")
    print_dict(regions)
    return

def print_menu(regions):
    print("-----------------------\n",
            "    menu:    ", 
            "\n-----------------------")
    print("option 1: add regions")
    print("option 2: remove regions")
    print("option 3: finished editing regions")
    return


def choose_option(regions):
    print_regions(regions)
    print_menu(regions)
    
    while True:
        option = input("Please enter menu option of your choice (integer): ")
        if option in ('1', '2', '3'):
            option = int(option)
            return option
        else:
            print("Your input '" + option + "' was invalid.", 
                    "Please choose '1', '2', or '3'.")
    return

def add_regions(regions):
    print_regions(regions)
    while True:
        try:
            num = int(input("Please enter the region number (integer): "))
        except ValueError:
            print("Your input was not an integer. Try again.")
            continue
        name = input("Please enter region name (must be same as excel sheet): ")
        if num in regions:
            overwrite = y_or_n("There is already a region with that number." +  
                                " Would you like to overwrite it?")
            if not overwrite:
                continue

        regions[num] = name
        print_regions(regions)
        finished = y_or_n("Would you like to add another region?")
        if not finished:
            break
    return


def remove_regions(regions):
    print_regions(regions)
    while True:
        try:
            num = int(input("Please enter the region number (integer) to delete: "))
        except ValueError:
            print("Your input was not an integer. Try again.")
            continue
        if num in regions:
            del regions[num]
            print_regions(regions)
            finished = y_or_n("Would you like to remove another region?")
            if not finished:
                return
        else:
            print("Your input '" + str(num) + "' is not a region.",
                    "Please choose again.")
    return
 

def sort_regions(regions, ws = None):
    if ws is not None:
        for row in ws.iter_rows():
            regions[row[0].value] = row[1].value

    if regions == {}:
        print("There is no sheet 'list' in workbook.")
        print("Please either add region and region numbers manually,",
                "or create a worksheet called 'list'.")
    
    print("Use this to add or remove regions.")
    while True:
        option = choose_option(regions)
        if option == 1:
           add_regions(regions)
           continue
        elif option == 2:
           remove_regions(regions)
           continue
        # check if finished, then return
        elif option == 3:
            while True:
                finished = y_or_n("Please confirm. Are all regions correct?")
                if finished:
                    return
                break

def change_csv_names(save_csv_names):
    while True:
        change = input("Which file name would you like to change?" +
                       " ('map', 'legend', or 'overlaps'): ")
        
        q_pt1 = "What would you like to change the saved file name for '"
        q_pt2 = "' to?"

        if change in ('map', "'map'", '"map"'):
            save_csv_names['map'] = answer(q_pt1 + change + q_pt2)
        elif change in ('legend', "'legend'", '"legend"'):
            save_csv_names['legend'] = answer(q_pt1 + change + q_pt2)
        elif change in ('overlaps', "'overlaps'", '"overlaps"'):
            save_csv_names['overlaps'] = answer(q_pt1 + change + q_pt2)
        else:
            print("Your answer was not understood.")

        print_dict(save_csv_names)
        another = y_or_n("Would you like to change another filename?")
        if not another:
            return

def print_explain_csv(save_csv_names):
    print("In the folder 'Outputs', " + 
          "these are the csv file names that the corresponding information will be saved to:") 
    print_dict(save_csv_names)
    print("***Please note that this program will overwrite existing files.***")


##############################################################################
# defining variables
##############################################################################


def define_variable(variable_name, ws=None):
    line_end = "===================="
    line_begin = "\n" + line_end

    if variable_name == "xlFilename":
        print(line_begin, "input excel file", line_end)
        xlFilename = "individual_region_files.xlsx"
        print('In the folder "Outputs", ' + 
              'the current input excel file name is "' + xlFilename +'".')
        change = y_or_n("Would you like to change the input file name")
        if change:
            xlFilename = answer("Please enter input excel filename")
        return xlFilename
    
    elif variable_name == "area_name":
        print(line_begin, "area name", line_end)
        area_name = answer("Please enter the name of the entire area" + 
                           " (must match the excel sheet name)")
        return area_name
    
    elif variable_name == "regions":
        regions = {}
        print(line_begin, "regions", line_end)
        print("Define region names", 
              "(region names should be the same as excel sheet names)")
        if ws is not None:
            sort_regions(regions,ws)
        else:
            sort_regions(regions)
        return regions
    
    elif variable_name == "save_csv_names":
        print(line_begin, "output CSV file names", line_end)
        save_csv_names = {
            'map'       :'map.csv',
            'legend'    :'legend.csv',
            'overlaps'  :'overlaps.csv'
        }
        print_explain_csv(save_csv_names)
        change = y_or_n("Would you like to change csv names")
        if change:
            change_csv_names(save_csv_names)
        return save_csv_names

    elif variable_name == "save_wb_name":
        print(line_begin, "output excel workbook name", line_end)
        save_wb_name = "formatted_map.xlsx"
        print('In the folder "Outputs", ' + 
              'the current input excel file name is "' + save_wb_name +'".')
        change = y_or_n("***Please note that this program will overwrite the existing file***" + 
                        "\nWould you like to change the output workbook file name?")
        if change:
            save_wb_name = answer("Please enter output workbook filename" +  
                                    " (must be different than input name)")
        return save_wb_name
    
    else:
        print("Problem: variable name does not exist")
        return('')

def load_input_workbook(xlFilename):
    print("Now loading workbook: " + xlFilename)
    wb = load_workbook(xlFilename)
    print("Finished loading workbook: " + xlFilename)
    return wb

def define_all_variables():

    # define variables
    xlFilename = define_variable("xlFilename")
    wb = load_input_workbook(xlFilename)

    save_csv_names = define_variable("save_csv_names")
    format_map = y_or_n("\nShould the regionalized map also be formatted" + 
                        " and saved as an excel workbook?")
    save_wb_name = ''
    if format_map:
        while True:
            save_wb_name = define_variable("save_wb_name")
            if xlFilename == save_wb_name:
                print("ERROR: filename to save to cannot be the same as original workbook")
            else:
                break

    while True:
        area_name = define_variable("area_name")
        if area_name not in wb.sheetnames:
            print("ERROR: No sheet with this name was found")
        else:
            break

    if 'list' not in wb.sheetnames:
        regions = define_variable("regions")
    else:
        regions = define_variable("regions", wb["list"])

    return xlFilename, wb, save_csv_names, format_map, save_wb_name, area_name, regions
